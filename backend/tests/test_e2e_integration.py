import json
import os
import pytest
from fastapi.testclient import TestClient
from app.main import app
from app.core.local_embeddings import LocalEmbeddingModel
from app.core.documents.generator import (
    DocumentGenerator,
    CustomsInvoiceSchema,
    InvoiceItemSchema,
    SupplyAgreementSchema,
)


@pytest.mark.asyncio
async def test_e2e_classification_calculation_generation(monkeypatch, tmp_path):
    """
    End-to-end integration test:
    1. Turn 1 (HS classification): Ask orchestrator to classify a product description (e.g. computer/notebook).
    2. Turn 2 (Customs calculation): Chain conversational history and request customs payment calculation.
    3. Document Generation: Use calculation outputs to generate formal Excel and Word documents via DocumentGenerator.
    """
    # 1. Setup/Mock
    monkeypatch.setattr(LocalEmbeddingModel, "_available", False)
    client = TestClient(app)

    # 2. Turn 1: Classify product description
    user_msg_1 = "Определи код ТН ВЭД для портативного ноутбука из Китая"
    resp1 = client.post("/api/orchestrate", data={"text": user_msg_1})
    assert resp1.status_code == 200
    data1 = resp1.json()
    assert data1["intent"] == "product_description"
    assert len(data1["message"]) > 0

    # We will simulate a standard assistant response detailing the selected code, duty rate, and recycling status
    assistant_msg_1 = "Рекомендуемый код ТН ВЭД: 8517130000. Пошлина 10.0%. Требуется уплатить утильсбор ♻."

    # 3. Turn 2: Calculation request with history chaining
    history = [
        {"role": "user", "content": user_msg_1},
        {"role": "assistant", "content": assistant_msg_1},
    ]

    user_msg_2 = "Посчитай таможенную пошлину для партии стоимостью $12500"
    resp2 = client.post(
        "/api/orchestrate", data={"text": user_msg_2, "history": json.dumps(history)}
    )
    assert resp2.status_code == 200
    data2 = resp2.json()
    assert data2["intent"] == "calculation_request"
    assert "Автоматический расчёт таможенных платежей" in data2["message"]

    # Verify context-chaining display warning is present
    assert data2["chain_warning"] is not None
    assert "8517130000" in data2["chain_warning"]
    assert "10.0" in data2["chain_warning"]
    assert "утильсбор" in data2["chain_warning"]

    # Verify structured calculation pipeline results
    results = data2.get("pipeline_results", {})
    assert "calculation_request" in results
    assert "calculation_response" in results

    calc_req = results["calculation_request"]
    calc_res = results["calculation_response"]

    assert calc_req["invoice_price"] == 12500.0
    assert calc_req["currency"] == "USD"
    assert calc_req["duty_rate_percent"] == 10.0
    assert calc_req["is_subject_to_recycling_fee"] is True

    # Verify calculated values are present and numeric
    assert calc_res["customs_value_kzt"] > 0
    assert calc_res["customs_duty_kzt"] > 0
    assert calc_res["import_vat_kzt"] > 0
    assert calc_res["recycling_fee_kzt"] > 0
    assert calc_res["total_payments_kzt"] > 0

    # 4. Phase 3: Formal trade document generation using the chained output
    # Prepare schemas matching the calculated results
    invoice_path = str(tmp_path / "customs_invoice.xlsx")
    contract_path = str(tmp_path / "supply_agreement.docx")

    invoice_schema = CustomsInvoiceSchema(
        seller_name="Shenzhen Tech Corp",
        buyer_name="SmartKeden Almaty LLP",
        incoterms="FCA Shenzhen",
        items=[
            InvoiceItemSchema(
                name="Портативный ноутбук (из классификатора)",
                hs_code="8517130000",
                qty=25.0,
                unit="pcs",
                price=500.0,  # 25 * 500 = 12500 USD total invoice price
            )
        ],
    )

    contract_schema = SupplyAgreementSchema(
        contract_no="SK-2026-CH-088",
        contract_date="28 мая 2026 г.",
        seller_name="Shenzhen Tech Corp",
        buyer_name="SmartKeden Almaty LLP",
        incoterms="FCA Shenzhen",
    )

    # Invoke DocumentGenerator to verify correct compilation and layout
    result_invoice_path = DocumentGenerator.generate_invoice_excel(
        invoice_schema, invoice_path
    )
    result_contract_path = DocumentGenerator.generate_contract_word(
        contract_schema, contract_path
    )

    assert os.path.exists(result_invoice_path)
    assert os.path.exists(result_contract_path)
    assert result_invoice_path == invoice_path
    assert result_contract_path == contract_path
    # 5. Validate generated file contents
    import openpyxl
    import docx as docx_lib

    wb = openpyxl.load_workbook(result_invoice_path)
    ws = wb["Commercial Invoice"]
    assert ws["A4"].value == "Shenzhen Tech Corp"
    assert ws["A9"].value == "SmartKeden Almaty LLP"
    doc = docx_lib.Document(result_contract_path)
    assert "ДОГОВОР ПОСТАВКИ №" in doc.paragraphs[0].text


@pytest.mark.asyncio
async def test_e2e_empty_request_returns_error():
    """Empty text in orchestrate endpoint should return a validation error."""
    client = TestClient(app)
    response = client.post("/api/orchestrate", data={"text": ""})
    assert response.status_code in (400, 422)
